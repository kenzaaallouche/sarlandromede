import streamlit as st
import sys
import datetime
import numpy as np
from PIL import Image
import pandas as pd
from st_aggrid import AgGrid, GridOptionsBuilder
from st_aggrid.shared import GridUpdateMode
from babel.dates import format_date, format_datetime, format_time
import openpyxl
from openpyxl import workbook,load_workbook,Workbook
from openpyxl.styles import Font, Fill
from openpyxl.formula.translate import Translator
from openpyxl.styles.borders import Border, Side
from io import BytesIO
import os

@st.cache_data()
def grat(file,option1,option2,option3,d):
    df =  pd.read_excel (file)
    df.columns= ['RegionNo', 'DistrictNo', 'DistrictNameE',
       'CityNo', 'AreaNo', 'SalesmanNo', 'SalesmanNameE', 'SalesmanWarehouse',
       'CustomerNo', 'CustomerNameE', 'InvoiceID', 'Date', 'ItemID',
       'ItemNameE', 'Qty', 'UOM', 'UnitPrice', 'FreeItem', 'PromotionDiscount',
       'LineValue', 'BUID','vide']
    df = df.drop(columns=['vide'])
    df=df.drop(['SalesmanNo','RegionNo', 'CityNo','AreaNo','DistrictNo', 'DistrictNameE', 'SalesmanNameE', 'CustomerNo','BUID', 'SalesmanWarehouse','UOM','UnitPrice', 'FreeItem','InvoiceID','Date',], axis=1)
    df = df.loc[df["PromotionDiscount"] != 0]
    
    dfg=df.groupby(['CustomerNameE','ItemID','ItemNameE'], as_index=False).sum()
    p=pd.pivot_table(dfg,index=["ItemID","ItemNameE"], 
                     columns=['CustomerNameE'], 
                     values=['PromotionDiscount'], aggfunc=np.sum)
    o=p.stack(dropna=True)
    o=o.unstack(level=2)
    if o.size !=0:
        o.to_excel ('e.xlsx')
        book=load_workbook('e.xlsx')
        sheet=book.active
        sheet.delete_rows(1,1)
        sheet.delete_rows(2,1)
        sheet['A1'].value='ItemID'
        book.save("e.xlsx")
        o=pd.read_excel ('e.xlsx')
        o=o.fillna(0)
        #o.columns
        myList =[]
        for i in list(o.columns):
            
            chars = ["'",',','.','!']
        
            res = i.translate(str.maketrans('', '', ''.join(chars)))
            myList.append(res)
        o.columns=myList
        myList =list(o.columns)
        #del myList[0:2]
        #del myList[len(myList)-1]
        #del myList[len(myList)-1]
        #del myList[len(myList)-1]
    
        t=pd.read_excel ('pro1.xlsx')
        o=t.merge(o, how='left', on='ItemID')
        o=o.fillna(0)
        book=load_workbook(option1+'.xlsx')
        for i in myList:
        # print(i)
            if i in book.sheetnames:
                book.active= book[i]
                sheet1=book.active
                for t in range(len(o['ItemID'])):
                    print(o['ItemID'])

                    sheet1['G'+str(t+12)].value=o[i][t]

                #print(o[i][t])
        nam=book.sheetnames
        for t in range(len(o['ItemID'])):
            book.active= book['BON DE PREPARATION']
            sheet1=book.active
            form='=+'
            for y in nam:
                if y=='BON DE PREPARATION' or y=='ItemNameE':
                    print ('non')
                elif y != nam[-1]:
                    form=form+"'"+y+"'"+"!G"+str(t+12)+"+"
                else:
                    form=form+"'"+y+"'"+"!G"+str(t+12)
                
            sheet1['G'+str(t+12)]=Translator(form, origin='G'+str(t+12)).translate_formula('G'+str(t+12))
        book.save(option1+'.xlsx') 


@st.cache_data()
def load_data(file,option1,option2,option3,d):
    df1 =  pd.read_excel (file)
    df1.columns= ['RegionNo', 'DistrictNo', 'DistrictNameE',
       'CityNo', 'AreaNo', 'SalesmanNo', 'SalesmanNameE', 'SalesmanWarehouse',
       'CustomerNo', 'CustomerNameE', 'InvoiceID', 'Date', 'ItemID',
       'ItemNameE', 'Qty', 'UOM', 'UnitPrice', 'FreeItem', 'PromotionDiscount',
       'LineValue', 'BUID','vide']
    df1 = df1.drop(columns=['vide'])
    df1=df1[df1['FreeItem'].isin(['Sold'])]
    maxc=len(df1[~df1.duplicated('CustomerNameE')]['CustomerNameE'])
    p=pd.pivot_table(df1, index=["ItemID","ItemNameE"], columns=['CustomerNameE'], values=['LineValue'], aggfunc=np.sum)
    p.to_excel('FG.xlsx')
    df =  pd.read_excel('FG.xlsx')
    m=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']
    print(maxc)
    book1=load_workbook('FG.xlsx')
    sheet1=book1.active
    sheet1.unmerge_cells('C1:'+m[maxc+1]+str('1'))
    for i, row in enumerate(sheet1):
        sheet2=row
    sheet1.delete_rows(1,1)
    sheet1.delete_rows(2,1)
    sheet1['A1'].value="ItemID"
    sheet1['B1'].value="ItemNameE"
    #sheet1.delete_rows(2,1)
    book1.save("book.xlsx")
    df =  pd.read_excel("book.xlsx")
    df=df.fillna(0)
    if option1 =='MM16F01':
        book=load_workbook('MM.xlsx')
    elif option1 =='SWS16F01':
        book=load_workbook('SWS.xlsx')
    else:
        book=load_workbook('PS.xlsx')
    a=book.sheetnames
    lis=[]
    for v in df.columns:
        
        chars = ["'",',','.','!']
     
        res = v.translate(str.maketrans('', '', ''.join(chars)))
        lis.append(res)
    for i in range(len(lis)): 
    
        ss_sheet1= book[a[i+1]]
        #print(lis[i+1])
        ss_sheet1.title =lis[i+1]
        ss_sheet1['B7'].value =lis[i+1]
        
        if i==len(lis)-2:
            break
    book.save("book1.xlsx")
    if option1 =='MM16F01':
        df2 =  pd.read_excel ("PRIXMM.xlsx")
    elif option1 =='SWS16F01':
        df2 =  pd.read_excel ("PRIXSWS.xlsx")
    else:
        df2 =  pd.read_excel ("PRIXPS.xlsx")
    df3=pd.read_excel ("book.xlsx")
    OP=df2.merge(df3, how='left', on='ItemID')
    OP["ct"]=OP["NBRUN"]*OP["PRIXUN"]
    OP=OP.fillna(0)
    myList =[]
    for i in list(OP.columns):
        
        chars = ["'",',','.','!']
     
        res = i.translate(str.maketrans('', '', ''.join(chars)))
        myList.append(res)
    OP.columns=myList
    del myList[0:5]
    del myList[len(myList)-1]
    for i in myList:
        OP[i]=OP[i]/OP["ct"]
    OP=OP.fillna(0)
    #book.sheetnames
    for i in myList:
        if i in book.sheetnames:
            book.active= book[i]
        
            sheet1=book.active
            for t in range(len(OP['ItemID'])):
                #print(t)
                sheet1['E'+str(t+12)].value=OP[i][t]
                
            #print(OP[i][t])
    book.save('fin.xlsx')   
    OP["total"]=0
    for i in myList:
        
        OP["total"]=OP["total"]+OP[i]
    nam=book.sheetnames
    for t in range(len(OP['ItemID'])):
        book.active= book['BON DE PREPARATION']
        sheet1=book.active
        form='=+'
        for y in nam:
            if y=='BON DE PREPARATION' or y=='ItemNameE':
                print ('non')
            elif y != nam[-1]:
                form=form+"'"+y+"'"+"!E"+str(t+12)+"+"
            else:
                form=form+"'"+y+"'"+"!E"+str(t+12)
            
        sheet1['E'+str(t+12)]=Translator(form, origin='E'+str(t+12)).translate_formula('E'+str(t+12))
    sheet1['B7']=option1+'-'+option2
    sheet1['B8']=option3
    sheet1['B9']=d
    book.save(option1+'.xlsx')   
hide_st_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            </style>
            """
st.markdown(hide_st_style, unsafe_allow_html=True)

st.title('BACK OFFICE SARL ANDROMED DISTRIBUTION -ALGER CENTRE- 2023')
st.text('Application Créée Par: ALLOUCHE KENZA')

st.session_state["Page1"]="Went Page1"

st.divider()
st.header('DATA BASE:')
uploaded_files = st.file_uploader("IMPORTER LE BON DE CHARGEMENT ", accept_multiple_files=True)
for uploaded_file in uploaded_files:
    bytes_data = uploaded_file.read()
    print (type(bytes_data))
    st.write("NOM DE FICHIER:", uploaded_file.name)
st.divider()
col1, col2,col3 = st.columns(3)


with col1:
    option1 = st.selectbox(
        "ROUTE :",
        ("MM16F01",
         "SWS16F01",
	 "WS16F01",
	 "WS16F02",
        "PS16F01",
         "PS16F02",
         "PS16F03",
         "PS16F04",
         "PS16F05",
         "PS16F06",
         "PS16F07",
         "PS16F08",
         "PS16F09",
         "PS16F10",
         "PS16F11",
         "PS16F12",
         "PS16F13",
         "PS16F14",
         "PS16F15",
         "PS16F16",
         "PS16F17",),
        key="v1",
    )

with col2:
    option2 = st.selectbox(
        "VENDEUR",
        
        ("MAHMOUDI MOHAMED",
	 "ISMAIL BOUZIANE",
	"YAHIAOUI  YOUCEF",
        "GUERRASSI HOUSSEM EDDINE",
        "GUERRASSI LOUAI",
        "LOULANSSA KHALED",
        "KADEM ISLAM",
        "ABDESSELAM FARID",
        "OUBOUCHOU KAMEL",
        "TABI ABDELRRAOUF",
        'SALHI FOUZI',
        "EL KADI ABDELMADJID MADJED",
        "AMMAM ABDELKRIM",
        "OUARTI YACINE",
        "DIAB ISMAIL",
        "HAMADACHE SOFIANE",
        "MANSOUR AHMED",
        "TOUADI  MOURAD",
	 'BENBOUZID AYMEN ',
	 'REZZOUG IMAD',
	 'MOHAMEDI MOKHTAR',
            	'OUAHIB ABDERRAHMANE',
		"SELMAN SEDDIK",
		'BOUREGAA YOUNES',
		'BELLALA MOHAMED',
            	'LAOUANA FOUAD',
            	'SAID HADJAZ',
            	'LAOUAR ZAKARIA',
            	'ALIOUA AYOUB',
            	'BEN TEFRAOUINE FAHIM',
            	'LEKBEDJ ABBES',
            	'ACILA ABDELLAH',
            	'BENNOUI HACHEM',
	 	'ADMANE IMAD',
	       
	    
	'BENSLIMENE BILLEL',
		 'HADJIEDJ ISLAM',
		 'MOUACI YACINE',
		 'TERKHACHE RIAD',
		 'MAMOUDI BILLAL RAMZI',
		 'HAOUAL SOUHIB',
		 'MIZI ALLAOUA ANIS',
		 'GUENDOUZ ABDERRAHMANE',
		 'SETIHI OUSSAMA',
		 'BOUGRA YOUNES',
	 'MEZIANE BOUDJEMAA',
		 'MECHAKOU RAFIK',
		'GHERDAOUI OUSSAMA',
		 'MEHTOUK KEDOUR',
	 
	

        ),
        key="v2",
        
    )
with col3:
    option3 = st.selectbox(
        "LIVREUR",
        (
		'VAN SUP',
            "MAHMOUDI MOHAMED",
	 "ISMAIL BOUZIANE",
	"YAHIAOUI  YOUCEF",
        "GUERRASSI HOUSSEM EDDINE",
        "GUERRASSI LOUAI",
        "LOULANSSA KHALED",
        "KADEM ISLAM",
        "ABDESSELAM FARID",
        "OUBOUCHOU KAMEL",
        "TABI ABDELRRAOUF",
        'SALHI FOUZI',
        "EL KADI ABDELMADJID MADJED",
        "AMMAM ABDELKRIM",
        "OUARTI YACINE",
        "DIAB ISMAIL",
        "HAMADACHE SOFIANE",
        "MANSOUR AHMED",
        "TOUADI  MOURAD",
	 'BENBOUZID AYMEN ',
	 'REZZOUG IMAD',
	 'MOHAMEDI MOKHTAR',
            	'OUAHIB ABDERRAHMANE',
		"SELMAN SEDDIK",
		'BOUREGAA YOUNES',
		'BELLALA MOHAMED',
            	'LAOUANA FOUAD',
            	'SAID HADJAZ',
            	'LAOUAR ZAKARIA',
            	'ALIOUA AYOUB',
            	'BEN TEFRAOUINE FAHIM',
            	'LEKBEDJ ABBES',
            	'ACILA ABDELLAH',
            	'BENNOUI HACHEM',
	 	'ADMANE IMAD',
	       
	    
	'BENSLIMENE BILLEL',
		 'HADJIEDJ ISLAM',
		 'MOUACI YACINE',
		 'TERKHACHE RIAD',
		 'MAMOUDI BILLAL RAMZI',
		 'HAOUAL SOUHIB',
		 'MIZI ALLAOUA ANIS',
		 'GUENDOUZ ABDERRAHMANE',
		 'SETIHI OUSSAMA',
		 'BOUGRA YOUNES',
	 'MEZIANE BOUDJEMAA',
			'MECHAKOU RAFIK',
		'GHERDAOUI OUSSAMA',
		 'MEHTOUK KEDOUR',


        ),
        key="v3",
        
    )
st.divider()

d = st.date_input("CHOISIR UNE DATE LIVRAISON",datetime.datetime.now())
format_date(d, locale='en')
st.write('DATE LIVRAISON:', d)
st.divider()

if st.button('EXECUTE'):
    

# Join various path components
    
    #fill=
    load_data(uploaded_file,option1,option2,option3,d)
    grat(uploaded_file,option1,option2,option3,d)
    
    with open(option1+'.xlsx', "rb") as template_file:
        template_byte = template_file.read()

    st.download_button(label="Click to Download Template File",
                        data=template_byte,
                        file_name=option1+'.xlsx',
                        )

    #fill=pd.ExcelWriter(option1+'.xlsx', engine='openpyxl')
    #fill =fill.to_excel(index=False).encode('utf-8')
    
    #st.download_button(label='📥 Download Current Result',
                                   # data=fill ,
                                    #file_name= 'df_test.xlsx',
                                   # mime="text/xlsx")




